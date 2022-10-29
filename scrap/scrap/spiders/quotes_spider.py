# import scrapy
# import openpyxl

# class QuotesSpider(scrapy.Spider):
#     name = "quotes"

#     def start_requests(self):

#             links = []
#             wookbook = openpyxl.load_workbook("1Comp.xlsx")
#             worksheet = wookbook.active
            
#             for i in range(1,10):
#                 for col in worksheet.iter_cols(8, 8):
#                     links.append(col[i].value)

#             for url in links:
#                 yield scrapy.Request(url=url, callback=self.parse)

#     def parse(self, response):
#         for quote in response.css('div.wrapper'):
#             yield {
#                 'text': quote.css('div.text').get(),
#             }

import scrapy
import openpyxl

class QuotesSpider(scrapy.Spider):
    name = "quotes"

    def start_requests(self):
        
        links = []
        wookbook = openpyxl.load_workbook("1Comp.xlsx")

        worksheet = wookbook.active

        for i in range(1,worksheet.max_row):
            for col in worksheet.iter_cols(8, 8):
                links.append(col[i].value)

        for url in links:
            yield scrapy.Request(url=url, callback=self.parse)

    def parse(self, response):
        page = response.url.split("/")[-2]
        filename = f'quotes-{page}.html'
        with open(filename, 'wb') as f:
            f.write(response.body)
        self.log(f'Saved file {filename}')